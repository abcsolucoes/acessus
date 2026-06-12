import argparse
import io
import os
import re
import time
import unicodedata
from collections import Counter
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook


PEOPLE_API = "https://people.googleapis.com/v1"
SHEET_NAME = "linhas_vivo"
DEFAULT_GROUP_NAME = "Linhas Corporativas Vivo"

IGNORAR_NOMES = {
    "DISPONIVEL",
    "DISPONIVEL RH",
    "VAGAS RH",
    "DISPONIVEL VER CHIP",
    "DISPONIVEL VER O CHIP",
    "DISPONIVEL CHIP",
    "DISPONIVEL CHIPS",
    "VAGA RH",
    "VAGAS",
}

PARTICULAS = {"da", "de", "do", "das", "dos", "e"}


def requests_client():
    try:
        import requests
    except ModuleNotFoundError as exc:
        raise RuntimeError("Dependencia ausente: instale com 'pip install requests openpyxl'.") from exc
    return requests


def carregar_env_file(path: str) -> None:
    with open(path, "r", encoding="utf-8") as env_file:
        for line in env_file:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


@dataclass
class LinhaVivo:
    nome_original: str
    nome_formatado: str
    telefone: str
    classificacao: str


def sem_acento(valor: str) -> str:
    normalized = unicodedata.normalize("NFD", valor or "")
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def chave(valor: str) -> str:
    return re.sub(r"\s+", " ", sem_acento(valor).strip()).upper()


def compactar(valor: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", chave(valor))


def normalizar_telefone(raw: Any) -> str:
    if raw is None:
        return ""

    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return ""

    if digits.startswith("015") and len(digits) == 14:
        return digits

    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]

    if len(digits) == 10:
        digits = digits[:2] + "9" + digits[2:]

    return "015" + digits


def palavras_significativas(nome_raw: str) -> list[str]:
    return [
        p
        for p in re.split(r"\s+", nome_raw.strip())
        if p and p.lower() not in PARTICULAS
    ]


def capitalizar(palavra: str) -> str:
    return palavra[:1].upper() + palavra[1:].lower()


def primeiro_ultimo(nome_raw: str) -> str:
    palavras = palavras_significativas(nome_raw)
    if not palavras:
        return ""
    if len(palavras) == 1:
        return capitalizar(palavras[0])
    return f"{capitalizar(palavras[0])} {capitalizar(palavras[-1])}"


def abreviar_nome(nome_raw: str) -> str:
    palavras = palavras_significativas(nome_raw)
    if len(palavras) < 3:
        return primeiro_ultimo(nome_raw)
    return f"{capitalizar(palavras[0])} {palavras[-2][0].upper()}. {capitalizar(palavras[-1])}"


def limpar_texto_classificacao(raw: Any) -> str:
    return re.sub(r"\s+", " ", str(raw or "").strip())


def normalizar_classificacao(raw: Any) -> str:
    """Converte o Departamento/Tipo da planilha no sufixo padrão do contato.

    Exemplos:
    - A - Metropolitana BH -> Equipe A
    - Supervisor -> Supervisor
    - Departamento Pessoal -> DP
    - Z - Exclusivos Vilma -> Exclusivo Vilma
    - Z - Exclusivos/MG -> Exclusivo
    - T- Vendedor externo -> Vendedor
    - Freelancer Equipe R -> Freelancer Equipe R
    """
    texto = limpar_texto_classificacao(raw)
    if not texto:
        return ""

    texto_chave = chave(texto)

    # Exclusivos: equipe Z não deve virar "Equipe Z".
    # Ela passa a ser classificada como Exclusivo + marca quando a marca estiver no texto.
    if texto_chave.startswith("Z") or "EXCLUSIVO" in texto_chave or "EXCLUSIVOS" in texto_chave:
        marcas = ("VILMA", "PORTO ALEGRE", "COGRAN", "CIMED", "INPROVETER", "LINDOYA")
        for marca in marcas:
            if marca in texto_chave:
                return "Exclusivo " + " ".join(capitalizar(p) for p in marca.lower().split())
        return "Exclusivo"

    # Equipes regionais: A, B, C... exceto Z, tratada acima.
    match_equipe = re.search(r"^\s*(?:EQUIPE\s*)?([A-Y])\s*(?:-|$)", texto_chave)
    if match_equipe:
        return f"Equipe {match_equipe.group(1)}"

    if "FREELANCER" in texto_chave:
        match = re.search(r"EQUIPE\s*([A-Z])", texto_chave) or re.search(r"\b([A-Z])\s*(?:-|$)", texto_chave)
        return f"Freelancer Equipe {match.group(1)}" if match else "Freelancer"

    mapas = [
        (("DEPARTAMENTO PESSOAL", "DP"), "DP"),
        (("RECURSOS HUMANOS", "RH"), "RH"),
        (("SUPERVISOR",), "Supervisor"),
        (("COMERCIAL",), "Comercial"),
        (("FINANCEIRO",), "Financeiro"),
        (("SUPORTE TECNICO", "SUPORTE"), "Suporte Técnico"),
        (("OPERACIONAL", "ESCRITORIO"), "Escritório"),
        (("VENDEDOR", "VENDEDOR EXTERNO"), "Vendedor"),
        (("PROMOTOR",), "Promotor"),
        (("CLIENTE",), "Cliente"),
        (("EXTERNO", "OUTROS"), "Externo / Outros"),
    ]

    for termos, classificacao in mapas:
        if any(termo in texto_chave for termo in termos):
            return classificacao

    # Fallback seguro: mantém o texto original capitalizado.
    return " ".join(capitalizar(p) if not p.isupper() else p for p in texto.split())


def montar_nome(nome_raw: str, classificacao: str, usar_abreviado: bool = False, salvar_como: str = "ABC") -> str:
    base = abreviar_nome(nome_raw) if usar_abreviado else primeiro_ultimo(nome_raw)
    salvar = "Particular" if chave(salvar_como) == "PARTICULAR" else "ABC"
    return f"{base} ({salvar}) - {classificacao}"


def deve_ignorar_nome(nome: str) -> bool:
    nome_chave = chave(nome)
    nome_compacto = compactar(nome)
    if not nome_chave:
        return True
    if nome_chave in IGNORAR_NOMES:
        return True
    return any(nome_compacto == compactar(item) for item in IGNORAR_NOMES)


def parece_corporativo(row: dict[str, Any]) -> bool:
    texto = " ".join(chave(str(v)) for v in row.values() if v is not None)
    if "PARTICULAR" in texto:
        return False
    return True


def achar_coluna(headers: list[str], candidatos: tuple[str, ...]) -> str | None:
    headers_por_chave = {compactar(h): h for h in headers}
    for candidato in candidatos:
        candidato_compacto = compactar(candidato)
        for h_compacto, original in headers_por_chave.items():
            if candidato_compacto in h_compacto or h_compacto in candidato_compacto:
                return original
    return None


def carregar_linhas_vivo(xlsx_bytes: bytes) -> list[LinhaVivo]:
    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index = None
    headers = []
    for idx, row in enumerate(rows[:20]):
        atuais = [str(c or "").strip() for c in row]
        compactos = [compactar(c) for c in atuais]
        tem_nome = any("NOME" in c or "COLABORADOR" in c or "USUARIO" in c for c in compactos)
        tem_numero = any("NUMERO" in c or "LINHA" in c or "TELEFONE" in c or "CELULAR" in c for c in compactos)
        if tem_nome and tem_numero:
            header_index = idx
            headers = atuais
            break

    if header_index is None:
        raise RuntimeError("Nao encontrei cabecalho com colunas de nome e telefone na aba linhas_vivo.")

    nome_col = achar_coluna(headers, ("nome", "colaborador", "usuario", "responsavel"))
    telefone_col = achar_coluna(headers, ("numero", "linha", "telefone", "celular"))
    classificacao_col = achar_coluna(headers, ("departamento", "equipe", "time", "setor", "classificacao", "tipo"))

    if not nome_col or not telefone_col or not classificacao_col:
        raise RuntimeError(
            "Nao consegui identificar as colunas obrigatorias. "
            f"Detectado: nome={nome_col}, telefone={telefone_col}, classificacao={classificacao_col}"
        )

    resultado: list[LinhaVivo] = []
    vistos: set[str] = set()

    for raw_row in rows[header_index + 1 :]:
        row = {
            headers[i]: raw_row[i] if i < len(raw_row) else None
            for i in range(len(headers))
            if headers[i]
        }
        nome = str(row.get(nome_col) or "").strip()
        classificacao = normalizar_classificacao(row.get(classificacao_col))
        telefone = normalizar_telefone(row.get(telefone_col))

        if deve_ignorar_nome(nome) or not classificacao or not telefone:
            continue
        if not parece_corporativo(row):
            continue

        dedupe_key = f"{chave(nome)}|{telefone}|{chave(classificacao)}"
        if dedupe_key in vistos:
            continue
        vistos.add(dedupe_key)

        resultado.append(
            LinhaVivo(
                nome_original=nome,
                nome_formatado=montar_nome(nome, classificacao),
                telefone=telefone,
                classificacao=classificacao,
            )
        )

    return resultado


def google_access_token() -> str:
    required = ("GOOGLE_CLIENT_ID", "GOOGLE_CLIENT_SECRET", "GOOGLE_REFRESH_TOKEN")
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        raise RuntimeError(f"Variaveis Google ausentes: {', '.join(missing)}")

    requests = requests_client()
    response = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": os.environ["GOOGLE_CLIENT_ID"],
            "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
            "refresh_token": os.environ["GOOGLE_REFRESH_TOKEN"],
            "grant_type": "refresh_token",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["access_token"]


def people_get(token: str, url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    requests = requests_client()
    response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, params=params, timeout=60)
    response.raise_for_status()
    return response.json()


def people_post(token: str, url: str, payload: dict[str, Any]) -> dict[str, Any]:
    requests = requests_client()
    response = request_com_retry(
        lambda: requests.post(url, headers={"Authorization": f"Bearer {token}"}, json=payload, timeout=60)
    )
    response.raise_for_status()
    return response.json() if response.content else {}


def people_patch(token: str, url: str, payload: dict[str, Any], params: dict[str, Any]) -> dict[str, Any]:
    requests = requests_client()
    response = request_com_retry(
        lambda: requests.patch(
            url,
            headers={"Authorization": f"Bearer {token}"},
            json=payload,
            params=params,
            timeout=60,
        )
    )
    if not response.ok:
        print(f"Erro HTTP {response.status_code} ao atualizar contato:")
        print(response.text)
    response.raise_for_status()
    return response.json()



def people_delete(token: str, resource_name: str) -> None:
    requests = requests_client()
    response = request_com_retry(
        lambda: requests.delete(
            f"{PEOPLE_API}/{resource_name}:deleteContact",
            headers={"Authorization": f"Bearer {token}"},
            timeout=60,
        )
    )
    if not response.ok:
        print(f"Erro HTTP {response.status_code} ao excluir contato:")
        print(response.text)
    response.raise_for_status()


def nome_tem_padrao_abc(nome: str) -> bool:
    """Retorna True para qualquer contato no padrão ABC.

    Aceita exemplos como:
    - Alessandra Barbosa (ABC) - DP
    - Camila Sebe (ABC) - Supervisor
    - Lincoln Silva (ABC) - Equipe V
    - Polliana S. Rocha (ABC) - Equipe E
    """
    nome = re.sub(r"\s+", " ", (nome or "").strip())
    return bool(re.search(r"\((ABC|PARTICULAR)\)\s*-\s*\S+", nome, flags=re.IGNORECASE))


def score_para_manter_contato(contato: dict[str, Any], nome_esperado: str = "") -> tuple[int, int, int, str]:
    """Quanto maior, melhor para manter.

    Prioridades:
    1. Nome exatamente igual ao esperado pela planilha, quando existir.
    2. Qualquer contato com padrão (ABC) - Setor, incluindo DP, Supervisor e Equipe.
    3. Nome com mais informação.
    """
    nome = nome_google(contato)
    nome_chave = chave(nome)
    esperado_chave = chave(nome_esperado)
    return (
        1 if esperado_chave and nome_chave == esperado_chave else 0,
        1 if nome_tem_padrao_abc(nome) else 0,
        len(nome or ""),
        nome_chave,
    )


def limpar_duplicados_google(
    contatos: list[dict[str, Any]],
    dry_run: bool,
    pausa: float,
    token: str,
    nomes_esperados_por_telefone: dict[str, str] | None = None,
) -> None:
    """Remove contatos duplicados pelo mesmo telefone.

    Fluxo:
    - Agrupa todos os contatos do Google pelo telefone normalizado.
    - Se houver duplicidade e pelo menos um contato estiver no padrão ABC,
      mantém o melhor contato padronizado e apaga os demais.
    - Se nenhum contato do grupo estiver no padrão ABC, não apaga nada.

    Isso permite limpar também DP, Supervisor, Comercial etc., mesmo quando esses
    setores não aparecem na aba linhas_vivo.
    """
    nomes_esperados_por_telefone = nomes_esperados_por_telefone or {}
    grupos: dict[str, list[dict[str, Any]]] = {}

    for contato in contatos:
        telefone = telefone_google(contato)
        if not telefone:
            continue
        grupos.setdefault(telefone, []).append(contato)

    removidos = 0
    grupos_sem_padrao = 0

    for telefone, lista in grupos.items():
        if len(lista) <= 1:
            continue

        esperado = nomes_esperados_por_telefone.get(telefone, "")
        padronizados = [c for c in lista if nome_tem_padrao_abc(nome_google(c))]

        if not padronizados:
            grupos_sem_padrao += 1
            print(f"\n[IGNORADO] Telefone {telefone} tem duplicados, mas nenhum contato com padrão ABC.")
            for contato in lista:
                print(f"  - {nome_google(contato)}")
            continue

        manter = max(lista, key=lambda c: score_para_manter_contato(c, esperado))

        print(f"\n[DUPLICADO] Telefone {telefone}")
        if esperado:
            print(f"  Esperado pela planilha: {esperado}")
        print(f"  Manter: {nome_google(manter)}")

        for contato in lista:
            if contato.get("resourceName") == manter.get("resourceName"):
                continue

            rn = contato.get("resourceName")
            print(f"  Excluir: {nome_google(contato) or rn}")

            if not dry_run and rn:
                people_delete(token, rn)
                time.sleep(pausa)

            removidos += 1

    print(f"\nLimpeza final: {removidos} contatos duplicados removidos.")
    if grupos_sem_padrao:
        print(f"Ignorados: {grupos_sem_padrao} grupos duplicados sem nenhum contato padronizado ABC.")

def request_com_retry(call, tentativas: int = 5):
    espera = 1.0
    response = None
    for tentativa in range(1, tentativas + 1):
        response = call()
        if response.status_code not in (429, 500, 502, 503, 504):
            return response
        if tentativa == tentativas:
            return response
        retry_after = response.headers.get("Retry-After")
        pausa = float(retry_after) if retry_after and retry_after.isdigit() else espera
        print(f"Rate limit/erro temporario HTTP {response.status_code}. Tentando novamente em {pausa:.1f}s...")
        time.sleep(pausa)
        espera *= 2
    return response


def listar_google_contatos(token: str) -> list[dict[str, Any]]:
    todos = []
    page_token = None
    while True:
        params = {"personFields": "names,phoneNumbers,emailAddresses,metadata", "pageSize": 1000}
        if page_token:
            params["pageToken"] = page_token
        data = people_get(token, f"{PEOPLE_API}/people/me/connections", params)
        todos.extend(data.get("connections", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            return todos


def nome_google(person: dict[str, Any]) -> str:
    names = person.get("names") or []
    if not names:
        return ""
    return names[0].get("displayName") or names[0].get("givenName") or ""


def telefone_google(person: dict[str, Any]) -> str:
    phones = person.get("phoneNumbers") or []
    if not phones:
        return ""
    return normalizar_telefone(phones[0].get("value") or "")


def payload_contato(nome: str, telefone: str, contato_atual: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = {
        "names": [{"givenName": nome}],
        "phoneNumbers": [{"value": telefone, "type": "mobile"}],
    }
    if contato_atual and contato_atual.get("etag"):
        payload["etag"] = contato_atual["etag"]
    if contato_atual and contato_atual.get("metadata"):
        payload["metadata"] = contato_atual["metadata"]
    return payload


def nomes_finais_da_planilha(linhas: list[LinhaVivo]) -> dict[str, str]:
    """Calcula o nome final esperado por telefone sem criar sufixos #2.

    O Google Contatos permite nomes repetidos. Por isso, não faz sentido colocar
    #2 no nome do contato. Quando duas pessoas geram o mesmo Nome Sobrenome na
    mesma equipe, usamos a versão abreviada; se ainda assim repetir, mantemos o
    mesmo nome sem sufixo.
    """
    contagem = Counter(chave(linha.nome_formatado) for linha in linhas)
    resultado: dict[str, str] = {}

    for linha in linhas:
        nome = linha.nome_formatado
        if contagem[chave(linha.nome_formatado)] > 1:
            nome = montar_nome(linha.nome_original, linha.classificacao, usar_abreviado=True)

        resultado[linha.telefone] = nome

    return resultado


def sincronizar(linhas: list[LinhaVivo], dry_run: bool, pausa: float) -> None:
    token = google_access_token()
    contatos = listar_google_contatos(token)

    por_telefone = {telefone_google(p): p for p in contatos if telefone_google(p)}
    nomes_por_telefone = nomes_finais_da_planilha(linhas)

    criados = 0
    atualizados = 0
    sem_mudanca = 0

    for linha in linhas:
        existente = por_telefone.get(linha.telefone)
        nome_atual = nome_google(existente) if existente else ""
        nome_final = nomes_por_telefone[linha.telefone]

        if existente:
            rn = existente["resourceName"]
            tel_atual = telefone_google(existente)
            if chave(nome_atual) == chave(nome_final) and tel_atual == linha.telefone:
                sem_mudanca += 1
                continue

            print(f"[ATUALIZAR] {nome_atual or rn} -> {nome_final} | {linha.telefone}")
            if not dry_run:
                people_patch(
                    token,
                    f"{PEOPLE_API}/{rn}:updateContact",
                    payload_contato(nome_final, linha.telefone, existente),
                    {"updatePersonFields": "names,phoneNumbers"},
                )
                time.sleep(pausa)
            atualizados += 1
        else:
            print(f"[CRIAR] {nome_final} | {linha.telefone}")
            if not dry_run:
                created = people_post(token, f"{PEOPLE_API}/people:createContact", payload_contato(nome_final, linha.telefone))
                por_telefone[linha.telefone] = created
                time.sleep(pausa)
            criados += 1
    print(f"\nResumo: {criados} criados, {atualizados} atualizados, {sem_mudanca} sem mudanca.")

    print("\nVerificando e apagando duplicados por telefone...")
    contatos_atualizados = listar_google_contatos(token)
    limpar_duplicados_google(
        contatos_atualizados,
        dry_run=dry_run,
        pausa=pausa,
        token=token,
        nomes_esperados_por_telefone=nomes_por_telefone,
    )

    if dry_run:
        print("DRY-RUN ativo: nenhum contato foi alterado. Use --apply para gravar no Google Contatos.")


def microsoft_access_token() -> str:
    required = ("MICROSOFT_TENANT_ID", "MICROSOFT_CLIENT_ID", "MICROSOFT_CLIENT_SECRET")
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        raise RuntimeError(f"Variaveis Microsoft ausentes: {', '.join(missing)}")

    requests = requests_client()
    response = requests.post(
        f"https://login.microsoftonline.com/{os.environ['MICROSOFT_TENANT_ID']}/oauth2/v2.0/token",
        data={
            "client_id": os.environ["MICROSOFT_CLIENT_ID"],
            "client_secret": os.environ["MICROSOFT_CLIENT_SECRET"],
            "grant_type": "client_credentials",
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["access_token"]


def baixar_linhas_vivo_sharepoint() -> bytes:
    requests = requests_client()
    token = microsoft_access_token()
    headers = {"Authorization": f"Bearer {token}"}

    site = requests.get(
        "https://graph.microsoft.com/v1.0/sites/solucoesabc.sharepoint.com:/sites/Documentos",
        headers=headers,
        timeout=60,
    )
    site.raise_for_status()
    site_id = site.json()["id"]

    drives = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives", headers=headers, timeout=60)
    drives.raise_for_status()
    drive_id = next(
        d["id"]
        for d in drives.json().get("value", [])
        if d.get("name") == "Dados ABC"
    )

    file_path = "TI-Suporte/Gestao de linhas e aparelhos/Gestao de linhas e aparelhos.xlsx"
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_path}:/content"
    file_response = requests.get(url, headers=headers, timeout=120)
    if file_response.status_code == 404:
        file_path = "TI-Suporte/Gestão de linhas e aparelhos/Gestão de linhas e aparelhos.xlsx"
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_path}:/content"
        file_response = requests.get(url, headers=headers, timeout=120)
    file_response.raise_for_status()
    return file_response.content


def main() -> None:
    parser = argparse.ArgumentParser(description="Padroniza contatos corporativos da aba linhas_vivo no Google Contatos.")
    parser.add_argument("--xlsx", help="Caminho para o arquivo Gestão de linhas e aparelhos.xlsx. Se omitido, baixa do SharePoint.")
    parser.add_argument("--env-file", help="Arquivo .env com GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET e GOOGLE_REFRESH_TOKEN.")
    parser.add_argument("--apply", action="store_true", help="Grava alteracoes no Google Contatos. Sem isso, roda em dry-run.")
    parser.add_argument("--pause", type=float, default=0.15, help="Pausa entre escritas na People API.")
    args = parser.parse_args()

    if args.env_file:
        carregar_env_file(args.env_file)
    elif os.path.exists(".env"):
        carregar_env_file(".env")

    if args.xlsx:
        with open(args.xlsx, "rb") as f:
            xlsx_bytes = f.read()
    else:
        xlsx_bytes = baixar_linhas_vivo_sharepoint()

    linhas = carregar_linhas_vivo(xlsx_bytes)
    print(f"{len(linhas)} linhas corporativas validas encontradas na aba {SHEET_NAME}.")
    sincronizar(linhas, dry_run=not args.apply, pausa=args.pause)


if __name__ == "__main__":
    main()
